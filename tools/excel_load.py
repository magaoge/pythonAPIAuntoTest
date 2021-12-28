# excel操作
from openpyxl import load_workbook
# column_index_from_string根据列名获取列索引
from openpyxl.utils import column_index_from_string

from tools.do_regx import DoRegx
from tools.mylog import MyLog
from tools.path_maneger import test_config_contrl_path, test_data_path
from tools.read_config import ReadConfig


class LoadExcel:

    def __init__(self, filePath, sheetName):
        self.filePath = filePath
        self.sheetName = sheetName
        self.file_screem = load_workbook(self.filePath)
        self.sheet = self.file_screem[self.sheetName]
        # # 获取表单所有的表单名
        # self.sheets = self.file_screem.sheetnames
        # 获取表单所有行
        self.maxRow = self.sheet.max_row
        # 获取表单所有列
        self.maxColumn = self.sheet.max_column

    def get_header(self):
        head = []
        for i in range(1, self.maxColumn + 1):
            head_value = self.sheet.cell(1, i).value
            head.append(head_value)
        return head

    # 定制化的，根据配置文件，将自己想要拉取的所有用例拉出来，进行测试，这里有一个矛盾点
    # 这里的方法定义（是否是静态）、配置读取、路径，都是可以提取出来的，但是如果提取出来，那以后每次调用都需要重新去读一次
    # 配置，在后续的引用中也不是很方便，当然也可以把读取配置的行为提取出来，但是总之很麻烦
    @staticmethod
    def load_excel(test_config_contrl_path,section,option):

        mode = eval(ReadConfig().readConfig(test_config_contrl_path, section, option))

        # 这里取出来的是需要参数化的
        init_data_dict = LoadExcel(test_data_path, "init").replace_dict()

        # 首先声明一个列表，用来承接所有的数据
        sheet_data = []
        # 循环遍历字典中的所有key,其中的key是表单名
        for sheetName in mode:
            dt = LoadExcel(test_data_path, sheetName)
            # 从head里取出来，每一列的列头信息
            head = dt.get_header()
            print(mode[sheetName])
            if mode[sheetName] == 'all':
                # 声明一个字典，用来承接每一行每一列的对应值
                # 从每一行中取出每一列的信息
                # 将每个列头于每一行的每一列的值，组合起来放入字典中
                for i in range(2, dt.maxRow + 1):
                    row_data = {}
                    for j in range(1, dt.maxColumn + 1):
                        every_row_column_value = dt.sheet.cell(i, j).value
                        # 如果检测到列名是data,那么就去检测data对应该单元格中信息是否包含${}信息，如果包含就取出来，然后
                        # 再取出来{}中具体的变量名，然后再对该字符串中的变量名进行替换
                        # 注意！！！！最好数据做好之后，现对数据提取转换进行测试
                        # MyLog().info("每一行的数据是：{}".format(every_row_column_value))
                        if head[j - 1] == "data" and every_row_column_value.find("${username}") != -1:
                            every_row_column_value = DoRegx.do_regx("\$\{(.*?)}", every_row_column_value, init_data_dict,row_data["caseId"])

                        # 列表索引从0开始，j需要-1
                        row_data[head[j - 1]] = every_row_column_value
                        # 为了方便后续的数据回写，可以根据用例所在的表单，回写到相应的表单
                        # 中，我们在行数据存储中，写入表单名，对用例进行管理
                        row_data["sheet_name"] = sheetName
                    sheet_data.append(row_data)
            else:
                for i in mode[sheetName]:
                    row_data = {}
                    for j in range(1, dt.maxColumn + 1):
                        every_row_column_value = dt.sheet.cell(i, j).value

                        if head[j - 1] == "data" and every_row_column_value.find("${username}") != -1:
                            every_row_column_value = DoRegx.do_regx("\$\{(.*?)}", every_row_column_value, init_data_dict,i)

                        # 列表索引从0开始，j需要-1
                        row_data[head[j - 1]] = every_row_column_value
                        row_data["sheet_name"] = sheetName
                    sheet_data.append(row_data)
        return sheet_data

    def replace_dict(self):
        head = self.get_header()
        init_data = {}
        for i in range(2, self.maxRow + 1):
            for j in range(1, self.maxColumn + 1):
                every_row_column_value = self.sheet.cell(i, j).value
                # 列表索引从0开始，j需要-1
                init_data[head[j - 1]] = every_row_column_value
                # 为了方便后续的数据回写，可以根据用例所在的表单，回写到相应的表单
                # 中，我们在行数据存储中，写入表单名，对用例进行管理
                init_data["sheet_name"] = self.sheetName
        return init_data


    @staticmethod
    # 数据回写添加表单名，进行指定表单的数据回写
    def write_back(sheetName ,rowNum, columName, value):
            dt = LoadExcel(test_data_path, sheetName)
            head = dt.get_header()
            column_index = head.index(columName) + 1
            dt.sheet.cell(rowNum, column_index).value = value
            dt.file_screem.save(dt.filePath)


if __name__ == '__main__':
    # 测试replace_data方法
    # init_data = LoadExcel(test_data_path, "init").replace_data()
    # print(init_data)

    # 测试load_excel方法
    data = LoadExcel.load_excel()
    print(data)
    print(len(data))
    # filePath = "../test_data/apiTestData.xlsx"
    # sheetName = "login"
    #
    # db = LoadExcel(filePath, sheetName)
    # db.get_header()
    #
    # data = db.load_excel()
    # print(data)

    # sheet_data = .loadExcel()
    # print(sheet_data)

# # 3.定位表单行列，python中直接从1开始索引定位
# res = sheet.cell(1,1).value
#
# # 数据回写，注意数据回写的时候要关闭Excel文件，不会报错，但是无法写入保存
# sheet.cell(1,5).value = "你好"
# # 必须要保存，不保存的话只是临时写入
# wb.save("../test_data/apiTestData.xlsx")
#
# res2 = sheet.cell(1,5).value
# print(res2)
#
# print(maxRow,maxRow)
